"""
Interactive Excel Budget Impact Model Generator.

Creates a fully formula-driven Excel workbook where:
- Users can modify input cells (yellow highlighting)
- All calculations update automatically via Excel formulas
- No Python or macros required for recalculation

This is designed for face-to-face payer discussions where
the user needs to run scenarios in real-time.
"""

from typing import Dict, List, Optional
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment, NamedStyle, Protection
    )
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .inputs import BIMInputs, UptakeScenario


class InteractiveExcelGenerator:
    """
    Generates a fully interactive Excel workbook for BIM.
    All calculations are done with Excel formulas.
    """

    # Style constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    CALC_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, inputs: BIMInputs):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        self.inputs = inputs
        self.wb = Workbook()
        self._setup_styles()

    def _setup_styles(self):
        """Set up named styles."""
        header_style = NamedStyle(name="header_style")
        header_style.font = self.HEADER_FONT
        header_style.fill = self.HEADER_FILL
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = self.BORDER
        self.wb.add_named_style(header_style)

        input_style = NamedStyle(name="input_style")
        input_style.fill = self.INPUT_FILL
        input_style.border = self.BORDER
        input_style.alignment = Alignment(horizontal="right")
        input_style.protection = Protection(locked=False)
        self.wb.add_named_style(input_style)

        result_style = NamedStyle(name="result_style")
        result_style.fill = self.RESULT_FILL
        result_style.border = self.BORDER
        result_style.alignment = Alignment(horizontal="right")
        result_style.font = Font(bold=True)
        self.wb.add_named_style(result_style)

        calc_style = NamedStyle(name="calc_style")
        calc_style.fill = self.CALC_FILL
        calc_style.border = self.BORDER
        calc_style.alignment = Alignment(horizontal="right")
        self.wb.add_named_style(calc_style)

    def generate(self, output_path: str) -> str:
        """Generate the interactive Excel workbook."""
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        self._create_inputs_sheet()
        self._create_results_sheet()
        self._create_instructions_sheet()

        self.wb.active = self.wb["Inputs"]
        self.wb.save(output_path)
        return output_path

    def _create_inputs_sheet(self):
        """Create the main inputs sheet with editable cells and formulas."""
        ws = self.wb.create_sheet("Inputs", 0)

        currency = self.inputs.country.currency_symbol

        # Title
        ws.merge_cells("B2:G2")
        ws["B2"] = "IXA-001 BUDGET IMPACT MODEL - INTERACTIVE"
        ws["B2"].font = Font(size=18, bold=True, color="1F4E79")

        ws["B3"] = "Modify YELLOW cells - all calculations update automatically!"
        ws["B3"].font = Font(size=12, italic=True, color="E67300")

        # ========== POPULATION INPUTS (Rows 5-13) ==========
        row = 5
        ws[f"B{row}"] = "POPULATION INPUTS"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:D{row}")

        # Row 6: Total Population
        ws["B6"] = "Total Plan Population"
        ws["C6"] = self.inputs.population.total_population
        ws["C6"].style = "input_style"
        ws["C6"].number_format = "#,##0"
        ws["D6"] = "lives"

        # Row 7: Adult Proportion
        ws["B7"] = "Adult Proportion (18+)"
        ws["C7"] = self.inputs.population.adult_proportion
        ws["C7"].style = "input_style"
        ws["C7"].number_format = "0.0%"

        # Row 8: HTN Prevalence
        ws["B8"] = "Hypertension Prevalence"
        ws["C8"] = self.inputs.population.hypertension_prevalence
        ws["C8"].style = "input_style"
        ws["C8"].number_format = "0.0%"

        # Row 9: Resistant HTN
        ws["B9"] = "Resistant HTN (of HTN)"
        ws["C9"] = self.inputs.population.resistant_htn_proportion
        ws["C9"].style = "input_style"
        ws["C9"].number_format = "0.0%"

        # Row 10: Uncontrolled
        ws["B10"] = "Uncontrolled (of Resistant)"
        ws["C10"] = self.inputs.population.uncontrolled_proportion
        ws["C10"].style = "input_style"
        ws["C10"].number_format = "0.0%"

        # Row 11: Treatment-seeking
        ws["B11"] = "Treatment-Seeking Rate"
        ws["C11"] = self.inputs.population.treatment_seeking_rate
        ws["C11"].style = "input_style"
        ws["C11"].number_format = "0.0%"

        # Row 13: Calculated eligible patients (FORMULA)
        ws["B13"] = "ELIGIBLE PATIENTS"
        ws["B13"].font = Font(bold=True, size=12)
        ws["C13"] = "=ROUND(C6*C7*C8*C9*C10*C11,0)"
        ws["C13"].style = "result_style"
        ws["C13"].number_format = "#,##0"

        # ========== MARKET INPUTS (Rows 15-20) ==========
        ws["B15"] = "BASELINE MARKET SHARES"
        ws["B15"].style = "header_style"
        ws.merge_cells("B15:D15")

        ws["B16"] = "Spironolactone"
        ws["C16"] = self.inputs.market.baseline_spironolactone
        ws["C16"].style = "input_style"
        ws["C16"].number_format = "0.0%"

        ws["B17"] = "Other MRA (Eplerenone)"
        ws["C17"] = self.inputs.market.baseline_other_mra
        ws["C17"].style = "input_style"
        ws["C17"].number_format = "0.0%"

        ws["B18"] = "No 4th-line Treatment"
        ws["C18"] = self.inputs.market.baseline_no_4th_line
        ws["C18"].style = "input_style"
        ws["C18"].number_format = "0.0%"

        ws["B19"] = "Total (should = 100%)"
        ws["C19"] = "=C16+C17+C18"
        ws["C19"].number_format = "0.0%"
        ws["C19"].font = Font(italic=True)

        # ========== DISPLACEMENT (Rows 21-25) ==========
        ws["B21"] = "DISPLACEMENT ASSUMPTIONS"
        ws["B21"].style = "header_style"
        ws.merge_cells("B21:D21")

        ws["B22"] = "From Spironolactone"
        ws["C22"] = self.inputs.market.displacement_from_spironolactone
        ws["C22"].style = "input_style"
        ws["C22"].number_format = "0.0%"

        ws["B23"] = "From Other MRA"
        ws["C23"] = self.inputs.market.displacement_from_other_mra
        ws["C23"].style = "input_style"
        ws["C23"].number_format = "0.0%"

        ws["B24"] = "From Untreated (New Starts)"
        ws["C24"] = self.inputs.market.displacement_from_untreated
        ws["C24"].style = "input_style"
        ws["C24"].number_format = "0.0%"

        # ========== COST INPUTS (Rows 26-40) ==========
        ws["B26"] = "ANNUAL COSTS (Per Patient)"
        ws["B26"].style = "header_style"
        ws.merge_cells("B26:D26")

        cost_data = [
            (27, "IXA-001 Drug Cost", self.inputs.costs.ixa_001_annual),
            (28, "Spironolactone Drug Cost", self.inputs.costs.spironolactone_annual),
            (29, "Other MRA Drug Cost", self.inputs.costs.other_mra_annual),
            (30, "IXA-001 Monitoring", self.inputs.costs.monitoring_ixa_001),
            (31, "Spironolactone Monitoring", self.inputs.costs.monitoring_spironolactone),
            (32, "Other MRA Monitoring", self.inputs.costs.monitoring_other_mra),
            (33, "No Treatment Monitoring", self.inputs.costs.monitoring_no_treatment),
            (34, "Office Visits (All)", self.inputs.costs.office_visits_annual),
            (35, "IXA-001 AE Management", self.inputs.costs.ae_management_ixa_001),
            (36, "Spironolactone AE Management", self.inputs.costs.ae_management_spironolactone),
            (37, "Other MRA AE Management", self.inputs.costs.ae_management_other_mra),
            (38, "IXA-001 Avoided Events Offset", self.inputs.costs.avoided_events_ixa_001_annual),
            (39, "Spironolactone Avoided Events", self.inputs.costs.avoided_events_spironolactone_annual),
            (40, "Other MRA Avoided Events", self.inputs.costs.avoided_events_other_mra_annual),
        ]

        for r, label, value in cost_data:
            ws[f"B{r}"] = label
            ws[f"C{r}"] = value
            ws[f"C{r}"].style = "input_style"
            ws[f"C{r}"].number_format = f'"{currency}"#,##0'

        # ========== CALCULATED NET COSTS (Rows 42-46) ==========
        ws["B42"] = "NET ANNUAL COST PER PATIENT"
        ws["B42"].style = "header_style"
        ws.merge_cells("B42:D42")

        # IXA-001: Drug + Monitoring + Office + AE - Offset
        ws["B43"] = "IXA-001"
        ws["C43"] = "=C27+C30+C34+C35-C38"
        ws["C43"].style = "calc_style"
        ws["C43"].number_format = f'"{currency}"#,##0'

        # Spironolactone
        ws["B44"] = "Spironolactone"
        ws["C44"] = "=C28+C31+C34+C36-C39"
        ws["C44"].style = "calc_style"
        ws["C44"].number_format = f'"{currency}"#,##0'

        # Other MRA
        ws["B45"] = "Other MRA"
        ws["C45"] = "=C29+C32+C34+C37-C40"
        ws["C45"].style = "calc_style"
        ws["C45"].number_format = f'"{currency}"#,##0'

        # No Treatment
        ws["B46"] = "No Treatment"
        ws["C46"] = "=C33+C34"
        ws["C46"].style = "calc_style"
        ws["C46"].number_format = f'"{currency}"#,##0'

        # ========== UPTAKE SCENARIOS (Rows 48-54) ==========
        ws["B48"] = "IXA-001 UPTAKE BY YEAR"
        ws["B48"].style = "header_style"
        ws.merge_cells("B48:G48")

        headers = ["Scenario", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        for i, header in enumerate(headers):
            ws.cell(row=49, column=2+i, value=header).style = "header_style"

        # Conservative (Row 50)
        ws["B50"] = "Conservative"
        for i, uptake in enumerate(self.inputs.market.uptake_curves[UptakeScenario.CONSERVATIVE]):
            cell = ws.cell(row=50, column=3+i, value=uptake)
            cell.style = "input_style"
            cell.number_format = "0%"

        # Moderate (Row 51)
        ws["B51"] = "Moderate"
        for i, uptake in enumerate(self.inputs.market.uptake_curves[UptakeScenario.MODERATE]):
            cell = ws.cell(row=51, column=3+i, value=uptake)
            cell.style = "input_style"
            cell.number_format = "0%"

        # Optimistic (Row 52)
        ws["B52"] = "Optimistic"
        for i, uptake in enumerate(self.inputs.market.uptake_curves[UptakeScenario.OPTIMISTIC]):
            cell = ws.cell(row=52, column=3+i, value=uptake)
            cell.style = "input_style"
            cell.number_format = "0%"

        # Scenario Selector (Row 54)
        ws["B54"] = "SELECTED SCENARIO"
        ws["B54"].font = Font(bold=True, size=12, color="1F4E79")
        ws["C54"] = "Moderate"
        ws["C54"].style = "input_style"
        ws["C54"].font = Font(bold=True, size=12)

        # Add dropdown for scenario selection
        dv = DataValidation(type="list", formula1='"Conservative,Moderate,Optimistic"', allow_blank=False)
        dv.error = "Please select a valid scenario"
        dv.errorTitle = "Invalid Scenario"
        ws.add_data_validation(dv)
        dv.add(ws["C54"])

        # ========== ACTIVE UPTAKE (Row 56-57) - Formula lookup ==========
        ws["B56"] = "ACTIVE UPTAKE (from selected scenario)"
        ws["B56"].font = Font(bold=True, italic=True)

        # Create lookup formulas for each year based on selected scenario
        for i in range(5):
            col = get_column_letter(3 + i)
            # IF(C54="Conservative", Conservative row, IF(C54="Moderate", Moderate row, Optimistic row))
            formula = f'=IF($C$54="Conservative",{col}50,IF($C$54="Moderate",{col}51,{col}52))'
            ws.cell(row=57, column=3+i, value=formula)
            ws.cell(row=57, column=3+i).style = "calc_style"
            ws.cell(row=57, column=3+i).number_format = "0%"

        # Set column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10

    def _create_results_sheet(self):
        """Create the results dashboard with all calculations."""
        ws = self.wb.create_sheet("Results")

        currency = self.inputs.country.currency_symbol

        # Title
        ws.merge_cells("B2:I2")
        ws["B2"] = "BUDGET IMPACT RESULTS"
        ws["B2"].font = Font(size=18, bold=True, color="1F4E79")

        # Scenario display
        ws["B3"] = "Scenario:"
        ws["C3"] = "=Inputs!C54"
        ws["C3"].font = Font(bold=True, size=14, color="2E75B6")

        # ========== KEY METRICS (Rows 5-12) ==========
        ws["B5"] = "KEY METRICS"
        ws["B5"].style = "header_style"
        ws.merge_cells("B5:D5")

        ws["B6"] = "Eligible Patients"
        ws["C6"] = "=Inputs!C13"
        ws["C6"].style = "result_style"
        ws["C6"].number_format = "#,##0"

        ws["B7"] = "Cost per IXA-001 Patient"
        ws["C7"] = "=Inputs!C43"
        ws["C7"].style = "result_style"
        ws["C7"].number_format = f'"{currency}"#,##0'

        ws["B8"] = "Cost per Spiro Patient"
        ws["C8"] = "=Inputs!C44"
        ws["C8"].style = "result_style"
        ws["C8"].number_format = f'"{currency}"#,##0'

        ws["B9"] = "Incremental Cost/Patient"
        ws["C9"] = "=C7-C8"
        ws["C9"].style = "result_style"
        ws["C9"].number_format = f'"{currency}"#,##0'

        # ========== YEAR-BY-YEAR CALCULATIONS (Rows 14-22) ==========
        ws["B14"] = "YEAR-BY-YEAR BUDGET IMPACT"
        ws["B14"].style = "header_style"
        ws.merge_cells("B14:I14")

        # Headers
        headers = ["Year", "IXA Uptake", "IXA Pts", "Spiro Pts", "MRA Pts", "None Pts", "New World", "Current World", "Impact"]
        for i, header in enumerate(headers):
            ws.cell(row=15, column=2+i, value=header).style = "header_style"

        # Data rows for Years 1-5
        for year in range(1, 6):
            r = 15 + year
            col_uptake = get_column_letter(2 + year)  # C, D, E, F, G for years 1-5

            # Year
            ws.cell(row=r, column=2, value=year)

            # IXA-001 Uptake (from active uptake row)
            ws.cell(row=r, column=3, value=f"=Inputs!{col_uptake}57")
            ws.cell(row=r, column=3).number_format = "0%"

            # IXA-001 Patients
            ws.cell(row=r, column=4, value=f"=ROUND(Inputs!$C$13*C{r},0)")
            ws.cell(row=r, column=4).number_format = "#,##0"

            # Spiro Patients (baseline - displaced by IXA)
            ws.cell(row=r, column=5, value=f"=ROUND(Inputs!$C$13*(Inputs!$C$16-C{r}*Inputs!$C$22),0)")
            ws.cell(row=r, column=5).number_format = "#,##0"

            # Other MRA Patients
            ws.cell(row=r, column=6, value=f"=ROUND(Inputs!$C$13*(Inputs!$C$17-C{r}*Inputs!$C$23),0)")
            ws.cell(row=r, column=6).number_format = "#,##0"

            # No Treatment Patients
            ws.cell(row=r, column=7, value=f"=ROUND(Inputs!$C$13*(Inputs!$C$18-C{r}*Inputs!$C$24),0)")
            ws.cell(row=r, column=7).number_format = "#,##0"

            # New World Cost
            ws.cell(row=r, column=8, value=f"=D{r}*Inputs!$C$43+E{r}*Inputs!$C$44+F{r}*Inputs!$C$45+G{r}*Inputs!$C$46")
            ws.cell(row=r, column=8).number_format = f'"{currency}"#,##0'

            # Current World Cost (no IXA-001)
            ws.cell(row=r, column=9, value=f"=ROUND(Inputs!$C$13*Inputs!$C$16,0)*Inputs!$C$44+ROUND(Inputs!$C$13*Inputs!$C$17,0)*Inputs!$C$45+ROUND(Inputs!$C$13*Inputs!$C$18,0)*Inputs!$C$46")
            ws.cell(row=r, column=9).number_format = f'"{currency}"#,##0'

            # Budget Impact
            ws.cell(row=r, column=10, value=f"=H{r}-I{r}")
            ws.cell(row=r, column=10).style = "calc_style"
            ws.cell(row=r, column=10).number_format = f'"{currency}"#,##0'

        # Total row
        ws["B21"] = "TOTAL (5-Year)"
        ws["B21"].font = Font(bold=True)
        ws["J21"] = "=SUM(J16:J20)"
        ws["J21"].style = "result_style"
        ws["J21"].number_format = f'"{currency}"#,##0'

        # ========== SUMMARY METRICS (Rows 23-27) ==========
        ws["B23"] = "SUMMARY"
        ws["B23"].style = "header_style"
        ws.merge_cells("B23:D23")

        ws["B24"] = "5-Year Budget Impact"
        ws["C24"] = "=J21"
        ws["C24"].style = "result_style"
        ws["C24"].number_format = f'"{currency}"#,##0'

        ws["B25"] = "Average Annual Impact"
        ws["C25"] = "=J21/5"
        ws["C25"].style = "result_style"
        ws["C25"].number_format = f'"{currency}"#,##0'

        ws["B26"] = "Year 1 PMPM"
        ws["C26"] = "=J16/(Inputs!C6*12)"
        ws["C26"].style = "result_style"
        ws["C26"].number_format = f'"{currency}"0.00'

        ws["B27"] = "Year 5 PMPM"
        ws["C27"] = "=J20/(Inputs!C6*12)"
        ws["C27"].style = "result_style"
        ws["C27"].number_format = f'"{currency}"0.00'

        # ========== CHART ==========
        chart = BarChart()
        chart.title = "Annual Budget Impact"
        chart.y_axis.title = f"Budget Impact ({self.inputs.costs.currency})"
        chart.x_axis.title = "Year"
        chart.type = "col"
        chart.style = 10

        data = Reference(ws, min_col=10, min_row=15, max_col=10, max_row=20)
        cats = Reference(ws, min_col=2, min_row=16, max_col=2, max_row=20)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        chart.width = 12
        chart.height = 8
        ws.add_chart(chart, "B30")

        # Set column widths
        ws.column_dimensions["B"].width = 22
        for col in range(3, 11):
            ws.column_dimensions[get_column_letter(col)].width = 14

    def _create_instructions_sheet(self):
        """Create instructions sheet."""
        ws = self.wb.create_sheet("Instructions")

        ws.merge_cells("B2:G2")
        ws["B2"] = "HOW TO USE THIS MODEL"
        ws["B2"].font = Font(size=18, bold=True, color="1F4E79")

        instructions = [
            "",
            "QUICK START",
            "-" * 50,
            "1. Go to the 'Inputs' sheet",
            "2. Modify any YELLOW cell to change assumptions",
            "3. All calculations update AUTOMATICALLY",
            "4. View results on the 'Results' sheet",
            "",
            "NO BUTTONS OR MACROS NEEDED!",
            "Just change a yellow cell and watch the results update.",
            "",
            "WHAT YOU CAN CHANGE",
            "-" * 50,
            "Population: Plan size, HTN prevalence rates",
            "Market Shares: Baseline treatment distribution",
            "Displacement: Where IXA-001 patients come from",
            "Costs: Drug costs, monitoring, adverse events",
            "Uptake: Year-by-year market penetration",
            "Scenario: Use dropdown to switch scenarios",
            "",
            "TIPS FOR PAYER MEETINGS",
            "-" * 50,
            "1. Set plan population to match payer's membership",
            "2. Start with 'Moderate' scenario as base case",
            "3. Switch scenarios to show range of outcomes",
            "4. Adjust IXA-001 price to find break-even point",
            "5. Results sheet has chart for visual impact",
            "",
            "COLOR LEGEND",
            "-" * 50,
            "YELLOW = Editable input (change these!)",
            "GREEN = Key results/outputs",
            "BLUE = Calculated values (don't edit)",
            "DARK BLUE = Headers",
        ]

        for i, line in enumerate(instructions):
            ws[f"B{3+i}"] = line
            if line and not line.startswith("-") and "=" not in line:
                if line == line.upper():
                    ws[f"B{3+i}"].font = Font(bold=True, size=12, color="1F4E79")

        ws.column_dimensions["B"].width = 60


def generate_interactive_model(inputs: BIMInputs, output_path: str) -> str:
    """Generate an interactive Excel BIM model."""
    generator = InteractiveExcelGenerator(inputs)
    return generator.generate(output_path)
