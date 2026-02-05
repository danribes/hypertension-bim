"""
Excel Generator for Budget Impact Model.

Creates a user-friendly Excel workbook with:
- Input dashboard
- Results dashboard with visualizations
- Scenario comparisons
- Sensitivity analysis
"""

from typing import Dict, List, Optional
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment, NamedStyle
    )
    from openpyxl.chart import BarChart, LineChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .inputs import BIMInputs, UptakeScenario, COUNTRY_CONFIGS
from .calculator import BIMCalculator, BIMResults


class ExcelGenerator:
    """
    Generates user-friendly Excel workbook for BIM.
    """

    # Professional color palette
    PRIMARY_DARK = "1F4E79"      # Deep blue
    PRIMARY_MED = "2E75B6"       # Medium blue
    PRIMARY_LIGHT = "BDD7EE"     # Light blue
    ACCENT_GREEN = "C6EFCE"      # Success green
    ACCENT_GREEN_DARK = "006100" # Dark green text
    ACCENT_YELLOW = "FFEB9C"     # Warning yellow
    ACCENT_YELLOW_DARK = "9C5700" # Dark yellow text
    ACCENT_RED = "FFC7CE"        # Error red
    ACCENT_RED_DARK = "9C0006"   # Dark red text
    NEUTRAL_LIGHT = "F2F2F2"     # Light gray
    NEUTRAL_MED = "D9D9D9"       # Medium gray

    # Style constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    SUBHEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    HIGHLIGHT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    THICK_BORDER = Border(
        left=Side(style='medium', color='1F4E79'),
        right=Side(style='medium', color='1F4E79'),
        top=Side(style='medium', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79')
    )

    def __init__(self, inputs: BIMInputs, results: Optional[BIMResults] = None):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")

        self.inputs = inputs
        self.results = results
        self.wb = Workbook()
        self._setup_styles()

    def _setup_styles(self):
        """Set up named styles for the workbook."""
        # Header style (primary)
        header_style = NamedStyle(name="header_style")
        header_style.font = self.HEADER_FONT
        header_style.fill = self.HEADER_FILL
        header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_style.border = self.BORDER
        self.wb.add_named_style(header_style)

        # Subheader style
        subheader_style = NamedStyle(name="subheader_style")
        subheader_style.font = self.SUBHEADER_FONT
        subheader_style.fill = self.SUBHEADER_FILL
        subheader_style.alignment = Alignment(horizontal="center", vertical="center")
        subheader_style.border = self.BORDER
        self.wb.add_named_style(subheader_style)

        # Input style
        input_style = NamedStyle(name="input_style")
        input_style.fill = self.INPUT_FILL
        input_style.border = self.BORDER
        input_style.alignment = Alignment(horizontal="right")
        self.wb.add_named_style(input_style)

        # Result style
        result_style = NamedStyle(name="result_style")
        result_style.fill = self.RESULT_FILL
        result_style.border = self.BORDER
        result_style.alignment = Alignment(horizontal="right")
        result_style.font = Font(bold=True)
        self.wb.add_named_style(result_style)

        # Highlight style (positive outcome)
        highlight_style = NamedStyle(name="highlight_style")
        highlight_style.fill = self.HIGHLIGHT_FILL
        highlight_style.border = self.BORDER
        highlight_style.font = Font(bold=True, color=self.ACCENT_GREEN_DARK)
        highlight_style.alignment = Alignment(horizontal="right")
        self.wb.add_named_style(highlight_style)

        # Warning style
        warning_style = NamedStyle(name="warning_style")
        warning_style.fill = self.WARNING_FILL
        warning_style.border = self.BORDER
        warning_style.font = Font(bold=True, color=self.ACCENT_YELLOW_DARK)
        warning_style.alignment = Alignment(horizontal="right")
        self.wb.add_named_style(warning_style)

        # Error style
        error_style = NamedStyle(name="error_style")
        error_style.fill = self.ERROR_FILL
        error_style.border = self.BORDER
        error_style.font = Font(bold=True, color=self.ACCENT_RED_DARK)
        error_style.alignment = Alignment(horizontal="right")
        self.wb.add_named_style(error_style)

    def _apply_table_formatting(self, ws, start_row: int, end_row: int, num_cols: int, start_col: int = 1):
        """Apply consistent table formatting with alternating rows."""
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(start_col, start_col + num_cols):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.BORDER
                if row_idx == start_row:
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif (row_idx - start_row) % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL

    def _create_title(self, ws, title: str, row: int = 1, col: int = 1, span: int = 4):
        """Create a formatted title cell."""
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(bold=True, size=18, color=self.PRIMARY_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return row + 1

    def _create_subtitle(self, ws, subtitle: str, row: int, col: int = 1, span: int = 4):
        """Create a formatted subtitle cell."""
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
        cell = ws.cell(row=row, column=col, value=subtitle)
        cell.font = Font(bold=True, size=12, color=self.PRIMARY_MED)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return row + 1

    def _create_section_header(self, ws, header: str, row: int, col: int = 1, span: int = 4):
        """Create a formatted section header."""
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
        cell = ws.cell(row=row, column=col, value=header)
        cell.style = "header_style"
        return row + 1

    def generate(self, output_path: str) -> str:
        """
        Generate the complete Excel workbook.

        Args:
            output_path: Path to save the Excel file

        Returns:
            Path to the generated file
        """
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Create all sheets
        self._create_cover_sheet()
        self._create_input_dashboard()
        self._create_population_sheet()
        self._create_market_sheet()
        self._create_cost_sheet()
        self._create_results_dashboard()
        self._create_scenario_comparison()
        self._create_sensitivity_sheet()
        self._create_documentation_sheet()

        # Save workbook
        self.wb.save(output_path)
        return output_path

    def _create_cover_sheet(self):
        """Create cover/instructions sheet."""
        ws = self.wb.create_sheet("Cover", 0)

        # Header banner
        for col in range(2, 8):
            for row in range(2, 5):
                ws.cell(row=row, column=col).fill = self.HEADER_FILL

        # Title
        ws.merge_cells("B2:G4")
        ws["B2"] = "BUDGET IMPACT MODEL\nIXA-001 for Resistant Hypertension"
        ws["B2"].font = Font(size=22, bold=True, color="FFFFFF")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Version badge
        ws["B5"] = "v4.0"
        ws["B5"].font = Font(size=10, bold=True, color=self.PRIMARY_DARK)
        ws["B5"].fill = PatternFill(start_color=self.PRIMARY_LIGHT, end_color=self.PRIMARY_LIGHT, fill_type="solid")
        ws["B5"].alignment = Alignment(horizontal="center")

        # Model info section
        info_start = 7
        ws.merge_cells(f"B{info_start}:D{info_start}")
        ws[f"B{info_start}"] = "MODEL INFORMATION"
        ws[f"B{info_start}"].style = "header_style"

        info_items = [
            ("Product:", "IXA-001 (Selective Aldosterone Synthase Inhibitor)"),
            ("Indication:", "Uncontrolled Resistant Hypertension"),
            ("Sponsor:", "Atlantis Pharmaceuticals"),
            ("Model Version:", "4.0 (Dual-Branch Phenotyping)"),
            ("Date:", "February 2026"),
            ("Country:", self.inputs.country.country_name),
        ]

        for i, (label, value) in enumerate(info_items):
            row = info_start + 1 + i
            ws[f"B{row}"] = label
            ws[f"B{row}"].font = Font(bold=True, color=self.PRIMARY_DARK)
            ws[f"C{row}"] = value
            ws[f"B{row}"].border = self.BORDER
            ws[f"C{row}"].border = self.BORDER
            if i % 2 == 0:
                ws[f"B{row}"].fill = self.ALT_ROW_FILL
                ws[f"C{row}"].fill = self.ALT_ROW_FILL

        # Quick Start Guide
        guide_start = info_start + len(info_items) + 3
        ws.merge_cells(f"B{guide_start}:D{guide_start}")
        ws[f"B{guide_start}"] = "QUICK START GUIDE"
        ws[f"B{guide_start}"].style = "header_style"

        instructions = [
            ("1.", "Input Dashboard", "Modify key model inputs"),
            ("2.", "Population", "Review patient cascade"),
            ("3.", "Market Dynamics", "Review uptake assumptions"),
            ("4.", "Results Dashboard", "View budget impact summary"),
            ("5.", "Scenario Comparison", "Compare uptake scenarios"),
            ("6.", "Sensitivity", "Run sensitivity analysis"),
        ]

        for i, (num, sheet, desc) in enumerate(instructions):
            row = guide_start + 1 + i
            ws[f"B{row}"] = num
            ws[f"B{row}"].font = Font(bold=True, color=self.PRIMARY_MED)
            ws[f"C{row}"] = sheet
            ws[f"C{row}"].font = Font(bold=True)
            ws[f"D{row}"] = desc
            ws[f"D{row}"].font = Font(italic=True, color="666666")

        # Key outputs section
        outputs_start = guide_start + len(instructions) + 3
        ws.merge_cells(f"B{outputs_start}:D{outputs_start}")
        ws[f"B{outputs_start}"] = "KEY MODEL OUTPUTS"
        ws[f"B{outputs_start}"].style = "header_style"

        if self.results:
            currency = self.inputs.country.currency_symbol
            output_row = outputs_start + 1
            outputs = [
                ("Eligible Patients", f"{self.results.total_eligible_patients:,}"),
                ("5-Year Budget Impact", f"{currency}{self.results.total_budget_impact_5yr:,.0f}"),
                ("Year 1 PMPM", f"{currency}{self.results.pmpm_impact_year1:.2f}"),
                ("Year 5 PMPM", f"{currency}{self.results.pmpm_impact_year5:.2f}"),
            ]
            for i, (label, value) in enumerate(outputs):
                row = output_row + i
                ws[f"B{row}"] = label
                ws[f"B{row}"].font = Font(bold=True)
                ws[f"C{row}"] = value
                ws[f"C{row}"].style = "result_style"
                ws[f"B{row}"].border = self.BORDER
                ws[f"C{row}"].border = self.BORDER

        # Color legend
        legend_start = outputs_start + 7
        ws.merge_cells(f"B{legend_start}:D{legend_start}")
        ws[f"B{legend_start}"] = "COLOR LEGEND"
        ws[f"B{legend_start}"].style = "subheader_style"

        legend_items = [
            (self.INPUT_FILL, "Editable Input"),
            (self.RESULT_FILL, "Calculated Result"),
            (self.HIGHLIGHT_FILL, "Positive Outcome"),
            (self.WARNING_FILL, "Attention Required"),
        ]
        for i, (fill, desc) in enumerate(legend_items):
            row = legend_start + 1 + i
            ws[f"B{row}"].fill = fill
            ws[f"B{row}"].border = self.BORDER
            ws[f"C{row}"] = desc
            ws[f"C{row}"].font = Font(size=10)

        # Set column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 45
        ws.column_dimensions["D"].width = 30

        # Freeze pane
        ws.freeze_panes = "A6"

    def _create_input_dashboard(self):
        """Create the main input dashboard."""
        ws = self.wb.create_sheet("Input Dashboard")

        # Title
        ws.merge_cells("B2:F2")
        ws["B2"] = "INPUT DASHBOARD"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        # Instructions
        ws["B3"] = "Yellow cells can be modified. Press Enter after changes to update calculations."
        ws["B3"].font = Font(italic=True, size=10)

        # Population Inputs Section
        row = 5
        ws[f"B{row}"] = "POPULATION INPUTS"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:D{row}")

        pop_inputs = [
            ("Total Plan Population", self.inputs.population.total_population, "lives"),
            ("Adult Proportion (18+)", self.inputs.population.adult_proportion * 100, "%"),
            ("Hypertension Prevalence", self.inputs.population.hypertension_prevalence * 100, "%"),
            ("Resistant HTN (of HTN)", self.inputs.population.resistant_htn_proportion * 100, "%"),
            ("Uncontrolled (of Resistant)", self.inputs.population.uncontrolled_proportion * 100, "%"),
            ("Treatment-Seeking Rate", self.inputs.population.treatment_seeking_rate * 100, "%"),
        ]

        for i, (label, value, unit) in enumerate(pop_inputs):
            r = row + 1 + i
            ws[f"B{r}"] = label
            ws[f"C{r}"] = value
            ws[f"C{r}"].style = "input_style"
            ws[f"D{r}"] = unit

        # Calculated eligible patients
        r = row + len(pop_inputs) + 2
        ws[f"B{r}"] = "ELIGIBLE PATIENTS"
        ws[f"B{r}"].font = Font(bold=True)
        ws[f"C{r}"] = self.inputs.population.eligible_patients
        ws[f"C{r}"].style = "result_style"
        ws[f"C{r}"].number_format = "#,##0"

        # Market Inputs Section
        row = r + 3
        ws[f"B{row}"] = "MARKET INPUTS"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:D{row}")

        market_inputs = [
            ("Baseline - Spironolactone", self.inputs.market.baseline_spironolactone * 100, "%"),
            ("Baseline - Other MRA", self.inputs.market.baseline_other_mra * 100, "%"),
            ("Baseline - No 4th Line Tx", self.inputs.market.baseline_no_4th_line * 100, "%"),
        ]

        for i, (label, value, unit) in enumerate(market_inputs):
            r = row + 1 + i
            ws[f"B{r}"] = label
            ws[f"C{r}"] = value
            ws[f"C{r}"].style = "input_style"
            ws[f"D{r}"] = unit

        # Cost Inputs Section
        row = r + 3
        ws[f"B{row}"] = "COST INPUTS (Annual)"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:D{row}")

        currency = self.inputs.country.currency_symbol
        cost_inputs = [
            ("IXA-001 Drug Cost", self.inputs.costs.ixa_001_annual, currency),
            ("Spironolactone Drug Cost", self.inputs.costs.spironolactone_annual, currency),
            ("Other MRA Drug Cost", self.inputs.costs.other_mra_annual, currency),
            ("IXA-001 Monitoring", self.inputs.costs.monitoring_ixa_001, currency),
            ("Spironolactone Monitoring", self.inputs.costs.monitoring_spironolactone, currency),
            ("Office Visits (All)", self.inputs.costs.office_visits_annual, currency),
        ]

        for i, (label, value, unit) in enumerate(cost_inputs):
            r = row + 1 + i
            ws[f"B{r}"] = label
            ws[f"C{r}"] = value
            ws[f"C{r}"].style = "input_style"
            ws[f"C{r}"].number_format = f'"{unit}"#,##0'

        # Scenario Selection
        row = r + 3
        ws[f"B{row}"] = "ANALYSIS SETTINGS"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:D{row}")

        r = row + 1
        ws[f"B{r}"] = "Selected Scenario"
        ws[f"C{r}"] = self.inputs.selected_scenario.value.capitalize()
        ws[f"C{r}"].style = "input_style"

        r += 1
        ws[f"B{r}"] = "Time Horizon"
        ws[f"C{r}"] = self.inputs.time_horizon_years
        ws[f"C{r}"].style = "input_style"
        ws[f"D{r}"] = "years"

        r += 1
        ws[f"B{r}"] = "Include Event Offsets"
        ws[f"C{r}"] = "Yes" if self.inputs.include_event_offsets else "No"
        ws[f"C{r}"].style = "input_style"

        # Set column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 10

    def _create_population_sheet(self):
        """Create population cascade sheet."""
        ws = self.wb.create_sheet("Population")

        # Title
        ws.merge_cells("B2:E2")
        ws["B2"] = "POPULATION CASCADE"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        # Cascade table
        cascade = self.inputs.population.get_cascade()

        headers = ["Stage", "Patients", "% of Previous", "% of Total"]
        for i, header in enumerate(headers):
            ws.cell(row=4, column=i+2, value=header).style = "header_style"

        cascade_items = [
            ("Total Plan Population", cascade["total_population"], 100.0, 100.0),
            ("Adults (18+)", cascade["adults_18_plus"],
             cascade["adults_18_plus"] / cascade["total_population"] * 100,
             cascade["adults_18_plus"] / cascade["total_population"] * 100),
            ("With Hypertension", cascade["with_hypertension"],
             cascade["with_hypertension"] / cascade["adults_18_plus"] * 100 if cascade["adults_18_plus"] > 0 else 0,
             cascade["with_hypertension"] / cascade["total_population"] * 100),
            ("Resistant Hypertension", cascade["resistant_hypertension"],
             cascade["resistant_hypertension"] / cascade["with_hypertension"] * 100 if cascade["with_hypertension"] > 0 else 0,
             cascade["resistant_hypertension"] / cascade["total_population"] * 100),
            ("Uncontrolled Resistant", cascade["uncontrolled_resistant"],
             cascade["uncontrolled_resistant"] / cascade["resistant_hypertension"] * 100 if cascade["resistant_hypertension"] > 0 else 0,
             cascade["uncontrolled_resistant"] / cascade["total_population"] * 100),
            ("Eligible for Treatment", cascade["eligible_for_treatment"],
             cascade["eligible_for_treatment"] / cascade["uncontrolled_resistant"] * 100 if cascade["uncontrolled_resistant"] > 0 else 0,
             cascade["eligible_for_treatment"] / cascade["total_population"] * 100),
        ]

        for i, (stage, patients, pct_prev, pct_total) in enumerate(cascade_items):
            row = 5 + i
            ws.cell(row=row, column=2, value=stage)
            ws.cell(row=row, column=3, value=patients).number_format = "#,##0"
            ws.cell(row=row, column=4, value=pct_prev).number_format = "0.0%"
            ws.cell(row=row, column=5, value=pct_total).number_format = "0.00%"

        # Highlight eligible patients row
        for col in range(2, 6):
            ws.cell(row=10, column=col).style = "result_style"

        # Add funnel chart data reference
        ws["B13"] = "Data Sources"
        ws["B13"].font = Font(bold=True)
        ws["B14"] = "Adult proportion: US Census / Eurostat"
        ws["B15"] = "HTN prevalence: NHANES / ESC Guidelines"
        ws["B16"] = "Resistant HTN: Carey et al., Circulation 2018"

        # Set column widths
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15

    def _create_market_sheet(self):
        """Create market dynamics sheet."""
        ws = self.wb.create_sheet("Market Dynamics")

        # Title
        ws.merge_cells("B2:H2")
        ws["B2"] = "MARKET DYNAMICS & UPTAKE"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        # Baseline market shares
        ws["B4"] = "BASELINE MARKET SHARES (No IXA-001)"
        ws["B4"].style = "header_style"
        ws.merge_cells("B4:D4")

        baseline = [
            ("Spironolactone", self.inputs.market.baseline_spironolactone),
            ("Other MRA (Eplerenone)", self.inputs.market.baseline_other_mra),
            ("No 4th-line Treatment", self.inputs.market.baseline_no_4th_line),
        ]

        for i, (treatment, share) in enumerate(baseline):
            row = 5 + i
            ws.cell(row=row, column=2, value=treatment)
            ws.cell(row=row, column=3, value=share).number_format = "0%"

        # Displacement assumptions
        ws["B10"] = "DISPLACEMENT ASSUMPTIONS"
        ws["B10"].style = "header_style"
        ws.merge_cells("B10:D10")

        displacement = [
            ("From Spironolactone", self.inputs.market.displacement_from_spironolactone),
            ("From Other MRA", self.inputs.market.displacement_from_other_mra),
            ("New Treatment (Untreated)", self.inputs.market.displacement_from_untreated),
        ]

        for i, (source, pct) in enumerate(displacement):
            row = 11 + i
            ws.cell(row=row, column=2, value=source)
            ws.cell(row=row, column=3, value=pct).number_format = "0%"

        # Uptake curves
        ws["B16"] = "IXA-001 UPTAKE BY SCENARIO"
        ws["B16"].style = "header_style"
        ws.merge_cells("B16:G16")

        # Headers
        headers = ["Scenario", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        for i, header in enumerate(headers):
            ws.cell(row=17, column=i+2, value=header).style = "header_style"

        # Data
        scenarios = [
            (UptakeScenario.CONSERVATIVE, "Conservative"),
            (UptakeScenario.MODERATE, "Moderate"),
            (UptakeScenario.OPTIMISTIC, "Optimistic"),
        ]

        for i, (scenario, name) in enumerate(scenarios):
            row = 18 + i
            ws.cell(row=row, column=2, value=name)
            for j, uptake in enumerate(self.inputs.market.uptake_curves[scenario]):
                ws.cell(row=row, column=3+j, value=uptake).number_format = "0%"

        # Add line chart for uptake
        chart = LineChart()
        chart.title = "IXA-001 Market Uptake by Scenario"
        chart.y_axis.title = "Market Share"
        chart.x_axis.title = "Year"
        chart.style = 10

        # Data references
        data = Reference(ws, min_col=3, min_row=17, max_col=7, max_row=20)
        cats = Reference(ws, min_col=3, min_row=17, max_col=7, max_row=17)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        chart.width = 15
        chart.height = 10
        ws.add_chart(chart, "B24")

        # Set column widths
        for col in range(2, 8):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_cost_sheet(self):
        """Create cost inputs sheet."""
        ws = self.wb.create_sheet("Costs")

        currency = self.inputs.country.currency_symbol

        # Title
        ws.merge_cells("B2:F2")
        ws["B2"] = f"COST INPUTS ({self.inputs.costs.currency})"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        # Per-patient annual costs table
        ws["B4"] = "ANNUAL PER-PATIENT COSTS"
        ws["B4"].style = "header_style"
        ws.merge_cells("B4:F4")

        headers = ["Cost Component", "IXA-001", "Spironolactone", "Other MRA", "No Treatment"]
        for i, header in enumerate(headers):
            ws.cell(row=5, column=i+2, value=header).style = "header_style"

        cost_rows = [
            ("Drug Cost",
             self.inputs.costs.ixa_001_annual,
             self.inputs.costs.spironolactone_annual,
             self.inputs.costs.other_mra_annual,
             self.inputs.costs.no_treatment_annual),
            ("Monitoring",
             self.inputs.costs.monitoring_ixa_001,
             self.inputs.costs.monitoring_spironolactone,
             self.inputs.costs.monitoring_other_mra,
             self.inputs.costs.monitoring_no_treatment),
            ("Office Visits",
             self.inputs.costs.office_visits_annual,
             self.inputs.costs.office_visits_annual,
             self.inputs.costs.office_visits_annual,
             self.inputs.costs.office_visits_annual),
            ("AE Management",
             self.inputs.costs.ae_management_ixa_001,
             self.inputs.costs.ae_management_spironolactone,
             self.inputs.costs.ae_management_other_mra,
             self.inputs.costs.ae_management_no_treatment),
        ]

        for i, row_data in enumerate(cost_rows):
            row = 6 + i
            for j, value in enumerate(row_data):
                cell = ws.cell(row=row, column=j+2, value=value)
                if j > 0:
                    cell.number_format = f'"{currency}"#,##0'

        # Subtotal
        row = 10
        ws.cell(row=row, column=2, value="Subtotal").font = Font(bold=True)
        for col in range(3, 7):
            subtotal = sum(ws.cell(row=r, column=col).value or 0 for r in range(6, 10))
            ws.cell(row=row, column=col, value=subtotal).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=col).font = Font(bold=True)

        # Avoided events offset
        row = 12
        ws.cell(row=row, column=2, value="Avoided CV Events (Offset)")
        ws.cell(row=row, column=3, value=-self.inputs.costs.avoided_events_ixa_001_annual).number_format = f'"{currency}"#,##0'
        ws.cell(row=row, column=4, value=-self.inputs.costs.avoided_events_spironolactone_annual).number_format = f'"{currency}"#,##0'
        ws.cell(row=row, column=5, value=-self.inputs.costs.avoided_events_other_mra_annual).number_format = f'"{currency}"#,##0'
        ws.cell(row=row, column=6, value=0).number_format = f'"{currency}"#,##0'

        # Net cost
        row = 14
        ws.cell(row=row, column=2, value="NET ANNUAL COST").style = "result_style"
        treatments = ["ixa_001", "spironolactone", "other_mra", "no_treatment"]
        for i, treatment in enumerate(treatments):
            net_cost = self.inputs.costs.get_total_annual_cost(treatment, include_offsets=True)
            ws.cell(row=row, column=3+i, value=net_cost).style = "result_style"
            ws.cell(row=row, column=3+i).number_format = f'"{currency}"#,##0'

        # Cost comparison bar chart
        chart = BarChart()
        chart.title = "Net Annual Cost per Patient"
        chart.y_axis.title = f"Cost ({self.inputs.costs.currency})"
        chart.type = "col"
        chart.style = 10

        data = Reference(ws, min_col=3, min_row=14, max_col=6, max_row=14)
        cats = Reference(ws, min_col=3, min_row=5, max_col=6, max_row=5)
        chart.add_data(data)
        chart.set_categories(cats)
        chart.shape = 4

        chart.width = 12
        chart.height = 8
        ws.add_chart(chart, "B17")

        # Set column widths
        ws.column_dimensions["B"].width = 25
        for col in range(3, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_results_dashboard(self):
        """Create results dashboard with key outputs and charts."""
        ws = self.wb.create_sheet("Results Dashboard")

        if not self.results:
            ws["B2"] = "No results available. Run the BIM calculator first."
            return

        currency = self.inputs.country.currency_symbol

        # Header banner
        for col in range(2, 9):
            ws.cell(row=2, column=col).fill = self.HEADER_FILL

        # Title
        ws.merge_cells("B2:H2")
        ws["B2"] = "BUDGET IMPACT RESULTS"
        ws["B2"].font = Font(size=18, bold=True, color="FFFFFF")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B3:H3")
        ws["B3"] = f"Scenario: {self.results.scenario.value.capitalize()} | Country: {self.inputs.country.country_name}"
        ws["B3"].font = Font(size=11, italic=True, color=self.PRIMARY_MED)
        ws["B3"].alignment = Alignment(horizontal="center")

        # Key metrics box with styled cards
        ws.merge_cells("B5:D5")
        ws["B5"] = "KEY METRICS"
        ws["B5"].style = "header_style"

        metrics = [
            ("Eligible Patients", f"{self.results.total_eligible_patients:,}", False),
            ("5-Year Budget Impact", f"{currency}{self.results.total_budget_impact_5yr:,.0f}", True),
            ("Average Annual Impact", f"{currency}{self.results.average_annual_impact:,.0f}", False),
            ("Year 1 PMPM", f"{currency}{self.results.pmpm_impact_year1:.2f}", False),
            ("Year 5 PMPM", f"{currency}{self.results.pmpm_impact_year5:.2f}", False),
            ("Cost per IXA-001 Patient", f"{currency}{self.results.cost_per_ixa_patient:,.0f}", False),
            ("Incremental Cost/Patient", f"{currency}{self.results.incremental_cost_per_ixa_patient:,.0f}", False),
        ]

        for i, (label, value, highlight) in enumerate(metrics):
            row = 6 + i
            cell_label = ws.cell(row=row, column=2, value=label)
            cell_label.font = Font(bold=True, color=self.PRIMARY_DARK)
            cell_label.border = self.BORDER
            cell_value = ws.cell(row=row, column=3, value=value)
            cell_value.border = self.BORDER
            cell_value.alignment = Alignment(horizontal="right")
            if highlight:
                cell_value.style = "highlight_style"
                cell_value.font = Font(bold=True, size=12, color=self.ACCENT_GREEN_DARK)
            else:
                cell_value.style = "result_style"
            if i % 2 == 0:
                cell_label.fill = self.ALT_ROW_FILL
                if not highlight:
                    cell_value.fill = self.ALT_ROW_FILL

        # Year-by-year table
        ws.merge_cells("B15:G15")
        ws["B15"] = "YEAR-BY-YEAR BUDGET IMPACT"
        ws["B15"].style = "header_style"

        headers = ["Year", "IXA-001 Patients", "Budget - Current", "Budget - New", "Budget Impact", "PMPM"]
        for i, header in enumerate(headers):
            ws.cell(row=16, column=i+2, value=header).style = "header_style"

        for i, yr in enumerate(self.results.yearly_results):
            row = 17 + i
            ws.cell(row=row, column=2, value=yr.year).border = self.BORDER
            ws.cell(row=row, column=3, value=yr.patients_ixa_001).number_format = "#,##0"
            ws.cell(row=row, column=4, value=yr.cost_current_world).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=5, value=yr.cost_new_world).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=6, value=yr.budget_impact).number_format = f'"{currency}"#,##0'
            pmpm = yr.budget_impact / self.inputs.population.total_population / 12
            ws.cell(row=row, column=7, value=pmpm).number_format = f'"{currency}"0.00'
            # Apply alternating row colors and borders
            for col in range(2, 8):
                ws.cell(row=row, column=col).border = self.BORDER
                if i % 2 == 0:
                    ws.cell(row=row, column=col).fill = self.ALT_ROW_FILL

        # Total row with highlight
        row = 17 + len(self.results.yearly_results)
        for col in range(2, 8):
            ws.cell(row=row, column=col).fill = self.RESULT_FILL
            ws.cell(row=row, column=col).border = self.BORDER
            ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=2, value="TOTAL")
        ws.cell(row=row, column=6, value=self.results.total_budget_impact_5yr).number_format = f'"{currency}"#,##0'

        # Budget impact chart
        chart = BarChart()
        chart.title = "Annual Budget Impact"
        chart.y_axis.title = f"Budget Impact ({self.inputs.costs.currency})"
        chart.x_axis.title = "Year"
        chart.type = "col"
        chart.style = 10

        data = Reference(ws, min_col=6, min_row=16, max_col=6, max_row=16 + len(self.results.yearly_results))
        cats = Reference(ws, min_col=2, min_row=17, max_col=2, max_row=16 + len(self.results.yearly_results))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        chart.width = 12
        chart.height = 8
        ws.add_chart(chart, "I5")

        # Market share evolution chart
        ws["B25"] = "MARKET SHARE EVOLUTION"
        ws["B25"].style = "header_style"
        ws.merge_cells("B25:F25")

        share_headers = ["Year", "IXA-001", "Spironolactone", "Other MRA", "No Treatment"]
        for i, header in enumerate(share_headers):
            ws.cell(row=26, column=i+2, value=header).style = "header_style"

        for i, yr in enumerate(self.results.yearly_results):
            row = 27 + i
            ws.cell(row=row, column=2, value=yr.year)
            ws.cell(row=row, column=3, value=yr.share_ixa_001).number_format = "0%"
            ws.cell(row=row, column=4, value=yr.share_spironolactone).number_format = "0%"
            ws.cell(row=row, column=5, value=yr.share_other_mra).number_format = "0%"
            ws.cell(row=row, column=6, value=yr.share_no_treatment).number_format = "0%"

        # Set column widths
        ws.column_dimensions["B"].width = 20
        for col in range(3, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_scenario_comparison(self):
        """Create scenario comparison sheet."""
        ws = self.wb.create_sheet("Scenario Comparison")

        if not self.results:
            ws["B2"] = "No results available."
            return

        # Run all scenarios
        calculator = BIMCalculator(self.inputs)
        all_results = calculator.run_all_scenarios()

        currency = self.inputs.country.currency_symbol

        # Title
        ws.merge_cells("B2:F2")
        ws["B2"] = "SCENARIO COMPARISON"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        # Comparison table
        headers = ["Metric", "Conservative", "Moderate", "Optimistic"]
        for i, header in enumerate(headers):
            ws.cell(row=4, column=i+2, value=header).style = "header_style"

        metrics = [
            ("5-Year Budget Impact", "total_budget_impact_5yr", f'"{currency}"#,##0'),
            ("Average Annual Impact", "average_annual_impact", f'"{currency}"#,##0'),
            ("Year 1 PMPM", "pmpm_impact_year1", f'"{currency}"0.00'),
            ("Year 5 PMPM", "pmpm_impact_year5", f'"{currency}"0.00'),
            ("Year 5 IXA-001 Patients", None, "#,##0"),
        ]

        for i, (label, attr, fmt) in enumerate(metrics):
            row = 5 + i
            ws.cell(row=row, column=2, value=label)
            for j, scenario in enumerate([UptakeScenario.CONSERVATIVE, UptakeScenario.MODERATE, UptakeScenario.OPTIMISTIC]):
                res = all_results[scenario]
                if attr:
                    value = getattr(res, attr)
                else:
                    value = res.yearly_results[-1].patients_ixa_001 if res.yearly_results else 0
                ws.cell(row=row, column=3+j, value=value).number_format = fmt

        # Year-by-year comparison
        ws["B12"] = "YEAR-BY-YEAR BUDGET IMPACT BY SCENARIO"
        ws["B12"].style = "header_style"
        ws.merge_cells("B12:E12")

        headers = ["Year", "Conservative", "Moderate", "Optimistic"]
        for i, header in enumerate(headers):
            ws.cell(row=13, column=i+2, value=header).style = "header_style"

        for year in range(1, 6):
            row = 13 + year
            ws.cell(row=row, column=2, value=year)
            for j, scenario in enumerate([UptakeScenario.CONSERVATIVE, UptakeScenario.MODERATE, UptakeScenario.OPTIMISTIC]):
                res = all_results[scenario]
                if year <= len(res.yearly_results):
                    value = res.yearly_results[year-1].budget_impact
                else:
                    value = 0
                ws.cell(row=row, column=3+j, value=value).number_format = f'"{currency}"#,##0'

        # Comparison chart
        chart = LineChart()
        chart.title = "Budget Impact by Scenario"
        chart.y_axis.title = f"Annual Budget Impact ({self.inputs.costs.currency})"
        chart.x_axis.title = "Year"
        chart.style = 10

        data = Reference(ws, min_col=3, min_row=13, max_col=5, max_row=18)
        cats = Reference(ws, min_col=2, min_row=14, max_col=2, max_row=18)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        chart.width = 15
        chart.height = 10
        ws.add_chart(chart, "B21")

        # Set column widths
        ws.column_dimensions["B"].width = 25
        for col in range(3, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_sensitivity_sheet(self):
        """Create sensitivity analysis sheet."""
        ws = self.wb.create_sheet("Sensitivity")

        currency = self.inputs.country.currency_symbol

        # Title
        ws.merge_cells("B2:G2")
        ws["B2"] = "SENSITIVITY ANALYSIS"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        if not self.results:
            ws["B4"] = "No results available."
            return

        # One-way sensitivity - IXA-001 price
        calculator = BIMCalculator(self.inputs)
        base_price = self.inputs.costs.ixa_001_annual
        price_range = [base_price * mult for mult in [0.5, 0.75, 1.0, 1.25, 1.5]]

        ws["B4"] = "IXA-001 PRICE SENSITIVITY"
        ws["B4"].style = "header_style"
        ws.merge_cells("B4:D4")

        headers = ["Annual Price", "5-Year Impact", "PMPM Year 5"]
        for i, header in enumerate(headers):
            ws.cell(row=5, column=i+2, value=header).style = "header_style"

        price_results = calculator.sensitivity_analysis("ixa_001_annual", price_range)
        for i, result in enumerate(price_results):
            row = 6 + i
            ws.cell(row=row, column=2, value=result["value"]).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=3, value=result["budget_impact_5yr"]).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=4, value=result["pmpm_year5"]).number_format = f'"{currency}"0.00'

        # Resistant HTN proportion sensitivity
        ws["B13"] = "RESISTANT HTN PREVALENCE SENSITIVITY"
        ws["B13"].style = "header_style"
        ws.merge_cells("B13:D13")

        headers = ["Prevalence", "Eligible Patients", "5-Year Impact"]
        for i, header in enumerate(headers):
            ws.cell(row=14, column=i+2, value=header).style = "header_style"

        prev_range = [0.08, 0.10, 0.12, 0.14, 0.16]
        prev_results = calculator.sensitivity_analysis("resistant_htn_proportion", prev_range)
        for i, result in enumerate(prev_results):
            row = 15 + i
            ws.cell(row=row, column=2, value=result["value"]).number_format = "0%"
            # Recalculate eligible patients
            temp_inputs = BIMInputs()
            temp_inputs.population.resistant_htn_proportion = result["value"]
            ws.cell(row=row, column=3, value=temp_inputs.population.eligible_patients).number_format = "#,##0"
            ws.cell(row=row, column=4, value=result["budget_impact_5yr"]).number_format = f'"{currency}"#,##0'

        # Price threshold analysis
        ws["B22"] = "PRICE THRESHOLD ANALYSIS"
        ws["B22"].style = "header_style"
        ws.merge_cells("B22:D22")

        ws["B23"] = "Budget-neutral price (5-year impact = $0):"
        threshold_price = calculator.price_threshold_analysis(0)
        if threshold_price:
            ws["D23"] = threshold_price
            ws["D23"].number_format = f'"{currency}"#,##0'
            ws["D23"].style = "result_style"
        else:
            ws["D23"] = "N/A"

        # Set column widths
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

    def _create_documentation_sheet(self):
        """Create technical documentation sheet."""
        ws = self.wb.create_sheet("Documentation")

        # Title
        ws.merge_cells("B2:G2")
        ws["B2"] = "TECHNICAL DOCUMENTATION"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        # Model overview
        ws["B4"] = "MODEL OVERVIEW"
        ws["B4"].font = Font(size=12, bold=True)

        overview = [
            "This Budget Impact Model (BIM) estimates the financial impact of introducing",
            "IXA-001 to a healthcare plan's formulary for treating resistant hypertension.",
            "",
            "The model compares a 'Current World' (no IXA-001) to a 'New World' (with IXA-001)",
            "over a 5-year time horizon, calculating the incremental budget impact.",
        ]
        for i, line in enumerate(overview):
            ws[f"B{5+i}"] = line

        # Key assumptions
        ws["B12"] = "KEY ASSUMPTIONS"
        ws["B12"].font = Font(size=12, bold=True)

        assumptions = [
            "1. Population remains stable over the analysis period",
            "2. Treatment costs are applied for full year per patient",
            "3. IXA-001 displaces existing treatments according to specified proportions",
            "4. Avoided CV events are calculated from linked CEA model results",
            "5. No discounting applied (standard for BIM)",
            "6. Treatment adherence and discontinuation not explicitly modeled",
        ]
        for i, assumption in enumerate(assumptions):
            ws[f"B{13+i}"] = assumption

        # Data sources
        ws["B21"] = "DATA SOURCES"
        ws["B21"].font = Font(size=12, bold=True)

        sources = [
            ("Epidemiology", "NHANES, ESC Guidelines, Carey et al. Circulation 2018"),
            ("Drug Costs", "RED BOOK/WAC (US), BNF/MIMS (UK)"),
            ("Medical Costs", "CMS Fee Schedule (US), NHS Reference Costs (UK)"),
            ("Market Shares", "IQVIA, Symphony Health, Internal market research"),
            ("CV Event Rates", "Linked IXA-001 Cost-Effectiveness Model"),
        ]

        headers = ["Category", "Source"]
        for i, header in enumerate(headers):
            ws.cell(row=22, column=i+2, value=header).style = "header_style"

        for i, (cat, source) in enumerate(sources):
            row = 23 + i
            ws.cell(row=row, column=2, value=cat)
            ws.cell(row=row, column=3, value=source)

        # Limitations
        ws["B30"] = "LIMITATIONS"
        ws["B30"].font = Font(size=12, bold=True)

        limitations = [
            "- Model does not capture indirect costs (productivity loss)",
            "- Market dynamics may differ from assumed displacement patterns",
            "- Avoided event calculations based on clinical trial populations",
            "- Real-world adherence may differ from assumptions",
        ]
        for i, limitation in enumerate(limitations):
            ws[f"B{31+i}"] = limitation

        # Contact
        ws["B37"] = "MODEL CONTACT"
        ws["B37"].font = Font(size=12, bold=True)
        ws["B38"] = "For questions about this model, contact:"
        ws["B39"] = "Atlantis Pharmaceuticals - HEOR Team"

        # Set column widths
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 60
