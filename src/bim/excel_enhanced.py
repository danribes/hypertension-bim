"""
Enhanced Excel Generator for Budget Impact Model.

Creates a comprehensive Excel workbook with additional sheets:
- Tornado Diagram (one-way sensitivity)
- Subgroup Analysis
- 10-Year Projection
- Event Analysis
- PSA Results
"""

from typing import Dict, List, Optional
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment, NamedStyle
    )
    from openpyxl.chart import BarChart, LineChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .inputs import ExtendedBIMInputs, UptakeScenario, EventType, SubgroupType
from .calculator import (
    EnhancedBIMCalculator,
    ExtendedBIMResults,
    TornadoResult,
    PSAResults,
)
from .excel_generator import ExcelGenerator


class EnhancedExcelGenerator(ExcelGenerator):
    """
    Enhanced Excel generator with additional analysis sheets.

    Extends ExcelGenerator with:
    - Tornado diagram sheet
    - Subgroup analysis sheet
    - 10-year projection sheet
    - Event analysis sheet
    - PSA results sheet
    """

    # Additional style constants
    TORNADO_LOW_FILL = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    TORNADO_HIGH_FILL = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
    SUBGROUP_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    def __init__(
        self,
        inputs: ExtendedBIMInputs,
        results: Optional[ExtendedBIMResults] = None
    ):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel generation. "
                "Install with: pip install openpyxl"
            )

        self.extended_inputs = inputs
        self.extended_results = results

        # Initialize parent with base results
        base_results = results.base_results if results else None
        super().__init__(inputs, base_results)

    def generate(self, output_path: str) -> str:
        """
        Generate the complete enhanced Excel workbook.

        Args:
            output_path: Path to save the Excel file

        Returns:
            Path to the generated file
        """
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Create standard sheets from parent
        self._create_cover_sheet()
        self._create_input_dashboard()
        self._create_population_sheet()
        self._create_market_sheet()
        self._create_cost_sheet()
        self._create_results_dashboard()
        self._create_scenario_comparison()

        # Create enhanced sheets
        self._create_tornado_sheet()
        self._create_subgroup_sheet()
        self._create_extended_horizon_sheet()
        self._create_event_analysis_sheet()
        self._create_psa_sheet()

        # Standard documentation
        self._create_documentation_sheet()

        # Save workbook
        self.wb.save(output_path)
        return output_path

    def _create_tornado_sheet(self):
        """Create tornado diagram sheet with one-way sensitivity analysis."""
        ws = self.wb.create_sheet("Tornado Diagram")

        currency = self.extended_inputs.country.currency_symbol

        # Title
        ws.merge_cells("B2:H2")
        ws["B2"] = "ONE-WAY SENSITIVITY ANALYSIS (TORNADO DIAGRAM)"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        if not self.extended_results or not self.extended_results.tornado_results:
            ws["B4"] = "Run tornado analysis first using EnhancedBIMCalculator.run_tornado_analysis()"
            return

        tornado = self.extended_results.tornado_results

        # Get base impact
        base_impact = self.extended_results.base_results.total_budget_impact_5yr

        ws["B4"] = f"Base Case 5-Year Budget Impact: {currency}{base_impact:,.0f}"
        ws["B4"].font = Font(bold=True)

        # Data table
        headers = ["Parameter", "Low Value", "High Value", "Impact at Low", "Impact at High", "Range"]
        for i, header in enumerate(headers):
            ws.cell(row=6, column=i+2, value=header).style = "header_style"

        for i, result in enumerate(tornado):
            row = 7 + i
            ws.cell(row=row, column=2, value=result.parameter_label)
            ws.cell(row=row, column=3, value=result.low_value).number_format = "#,##0.00"
            ws.cell(row=row, column=4, value=result.high_value).number_format = "#,##0.00"
            ws.cell(row=row, column=5, value=result.impact_at_low).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=6, value=result.impact_at_high).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=7, value=result.impact_range).number_format = f'"{currency}"#,##0'

        # Create tornado chart data
        # For tornado, we need to show deviation from base
        chart_start_row = 7 + len(tornado) + 3
        ws.cell(row=chart_start_row-1, column=2, value="TORNADO CHART DATA").style = "header_style"
        ws.merge_cells(f"B{chart_start_row-1}:E{chart_start_row-1}")

        chart_headers = ["Parameter", "Low Deviation", "High Deviation"]
        for i, header in enumerate(chart_headers):
            ws.cell(row=chart_start_row, column=i+2, value=header).style = "header_style"

        for i, result in enumerate(tornado):
            row = chart_start_row + 1 + i
            ws.cell(row=row, column=2, value=result.parameter_label)
            # Deviation from base (negative for low impact, positive for high)
            low_dev = result.impact_at_low - base_impact
            high_dev = result.impact_at_high - base_impact
            ws.cell(row=row, column=3, value=low_dev).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=4, value=high_dev).number_format = f'"{currency}"#,##0'

        # Create horizontal bar chart (tornado style)
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Tornado Diagram - 5-Year Budget Impact Sensitivity"
        chart.y_axis.title = None
        chart.x_axis.title = f"Deviation from Base ({self.extended_inputs.costs.currency})"

        # Data references
        data_end_row = chart_start_row + len(tornado)
        data = Reference(ws, min_col=3, min_row=chart_start_row, max_col=4, max_row=data_end_row)
        cats = Reference(ws, min_col=2, min_row=chart_start_row+1, max_col=2, max_row=data_end_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4

        chart.width = 18
        chart.height = 12
        ws.add_chart(chart, "I6")

        # Set column widths
        ws.column_dimensions["B"].width = 30
        for col in range(3, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_subgroup_sheet(self):
        """Create subgroup analysis sheet."""
        ws = self.wb.create_sheet("Subgroup Analysis")

        currency = self.extended_inputs.country.currency_symbol

        # Title
        ws.merge_cells("B2:H2")
        ws["B2"] = "SUBGROUP ANALYSIS"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        if not self.extended_results or not self.extended_results.subgroup_results:
            ws["B4"] = "No subgroup analysis available. Enable subgroups in ExtendedBIMInputs."
            return

        current_row = 4

        for subgroup_type, results_list in self.extended_results.subgroup_results.items():
            # Section header
            ws.cell(row=current_row, column=2, value=f"{subgroup_type.value.upper()} SUBGROUPS")
            ws.cell(row=current_row, column=2).style = "header_style"
            ws.merge_cells(f"B{current_row}:G{current_row}")
            current_row += 1

            # Table headers
            headers = ["Subgroup", "Patients", "Proportion", "5-Year Impact", "Impact/Patient"]
            for i, header in enumerate(headers):
                ws.cell(row=current_row, column=i+2, value=header).style = "header_style"
            current_row += 1

            # Data rows
            for sg_result in results_list:
                ws.cell(row=current_row, column=2, value=sg_result.subgroup_name)
                ws.cell(row=current_row, column=3, value=sg_result.patients).number_format = "#,##0"
                ws.cell(row=current_row, column=4, value=sg_result.proportion).number_format = "0.0%"
                ws.cell(row=current_row, column=5, value=sg_result.budget_impact_5yr).number_format = f'"{currency}"#,##0'
                ws.cell(row=current_row, column=6, value=sg_result.budget_impact_per_patient).number_format = f'"{currency}"#,##0'
                current_row += 1

            # Add spacing
            current_row += 2

        # Add bar chart comparing subgroups
        if self.extended_results.subgroup_results:
            first_type = list(self.extended_results.subgroup_results.keys())[0]
            first_results = self.extended_results.subgroup_results[first_type]

            chart_row = current_row + 2
            ws.cell(row=chart_row-1, column=2, value="SUBGROUP COMPARISON CHART DATA")
            ws.cell(row=chart_row-1, column=2).font = Font(bold=True)

            for i, sg in enumerate(first_results):
                ws.cell(row=chart_row + i, column=2, value=sg.subgroup_name)
                ws.cell(row=chart_row + i, column=3, value=sg.budget_impact_5yr)

            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = f"5-Year Budget Impact by {first_type.value.title()}"
            chart.y_axis.title = f"Budget Impact ({self.extended_inputs.costs.currency})"

            data = Reference(ws, min_col=3, min_row=chart_row-1,
                           max_col=3, max_row=chart_row + len(first_results) - 1)
            cats = Reference(ws, min_col=2, min_row=chart_row,
                           max_col=2, max_row=chart_row + len(first_results) - 1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            chart.width = 12
            chart.height = 8
            ws.add_chart(chart, "I4")

        # Set column widths
        ws.column_dimensions["B"].width = 30
        for col in range(3, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_extended_horizon_sheet(self):
        """Create 10-year projection sheet."""
        ws = self.wb.create_sheet("10-Year Projection")

        currency = self.extended_inputs.country.currency_symbol

        # Title
        ws.merge_cells("B2:H2")
        ws["B2"] = "10-YEAR BUDGET IMPACT PROJECTION"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        ws["B3"] = f"Note: Years 6-10 assume market share plateaus at Year 5 levels"
        ws["B3"].font = Font(italic=True, size=10)

        if not self.extended_results or not self.extended_results.extended_yearly_results:
            ws["B5"] = "Extended horizon calculation not available."
            return

        # Key metrics
        ws["B5"] = "SUMMARY METRICS"
        ws["B5"].style = "header_style"
        ws.merge_cells("B5:D5")

        metrics = [
            ("5-Year Budget Impact",
             sum(yr.budget_impact for yr in self.extended_results.extended_yearly_results[:5])),
            ("10-Year Budget Impact", self.extended_results.extended_total_impact),
            ("Average Annual Impact (10yr)",
             self.extended_results.extended_total_impact / 10),
        ]

        for i, (label, value) in enumerate(metrics):
            row = 6 + i
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=3, value=value).style = "result_style"
            ws.cell(row=row, column=3).number_format = f'"{currency}"#,##0'

        # Year-by-year table
        ws["B11"] = "YEAR-BY-YEAR PROJECTION"
        ws["B11"].style = "header_style"
        ws.merge_cells("B11:H11")

        headers = ["Year", "IXA-001 Share", "IXA-001 Patients", "Current World",
                   "New World", "Budget Impact", "Cumulative"]
        for i, header in enumerate(headers):
            ws.cell(row=12, column=i+2, value=header).style = "header_style"

        cumulative = 0
        for i, yr in enumerate(self.extended_results.extended_yearly_results):
            row = 13 + i
            cumulative += yr.budget_impact

            ws.cell(row=row, column=2, value=yr.year)
            ws.cell(row=row, column=3, value=yr.share_ixa_001).number_format = "0%"
            ws.cell(row=row, column=4, value=yr.patients_ixa_001).number_format = "#,##0"
            ws.cell(row=row, column=5, value=yr.cost_current_world).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=6, value=yr.cost_new_world).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=7, value=yr.budget_impact).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=8, value=cumulative).number_format = f'"{currency}"#,##0'

            # Highlight plateau years
            if yr.year > 5:
                for col in range(2, 9):
                    ws.cell(row=row, column=col).fill = self.SUBGROUP_FILL

        # Total row
        total_row = 13 + len(self.extended_results.extended_yearly_results)
        ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=7,
                value=self.extended_results.extended_total_impact).style = "result_style"
        ws.cell(row=total_row, column=7).number_format = f'"{currency}"#,##0'

        # Line chart for 10-year trend
        chart = LineChart()
        chart.title = "10-Year Budget Impact Projection"
        chart.y_axis.title = f"Annual Impact ({self.extended_inputs.costs.currency})"
        chart.x_axis.title = "Year"
        chart.style = 10

        data = Reference(ws, min_col=7, min_row=12, max_col=7, max_row=total_row-1)
        cats = Reference(ws, min_col=2, min_row=13, max_col=2, max_row=total_row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        chart.width = 15
        chart.height = 10
        ws.add_chart(chart, "J5")

        # Set column widths
        ws.column_dimensions["B"].width = 12
        for col in range(3, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_event_analysis_sheet(self):
        """Create event analysis sheet with rates, costs, and events avoided."""
        ws = self.wb.create_sheet("Event Analysis")

        currency = self.extended_inputs.country.currency_symbol

        # Title
        ws.merge_cells("B2:H2")
        ws["B2"] = "CLINICAL EVENT ANALYSIS"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        # Event rates table
        ws["B4"] = "EVENT RATES (per 1,000 patient-years)"
        ws["B4"].style = "header_style"
        ws.merge_cells("B4:F4")

        headers = ["Event Type", "IXA-001", "Spironolactone", "Other MRA", "No Treatment"]
        for i, header in enumerate(headers):
            ws.cell(row=5, column=i+2, value=header).style = "header_style"

        rates = self.extended_inputs.event_rates
        event_labels = {
            EventType.STROKE: "Stroke",
            EventType.MI: "Myocardial Infarction",
            EventType.HF: "Heart Failure Hospitalization",
            EventType.CKD: "CKD Progression",
            EventType.ESRD: "End-Stage Renal Disease",
            EventType.CV_DEATH: "CV Death",
            EventType.ALL_CAUSE_DEATH: "All-Cause Death",
        }

        row = 6
        for event in EventType:
            ws.cell(row=row, column=2, value=event_labels[event])
            ws.cell(row=row, column=3, value=rates.get_rate(event, "ixa_001")).number_format = "0.0"
            ws.cell(row=row, column=4, value=rates.get_rate(event, "spironolactone")).number_format = "0.0"
            ws.cell(row=row, column=5, value=rates.get_rate(event, "other_mra")).number_format = "0.0"
            ws.cell(row=row, column=6, value=rates.get_rate(event, "no_treatment")).number_format = "0.0"
            row += 1

        # Event costs table
        row += 2
        ws.cell(row=row, column=2, value="EVENT COSTS")
        ws.cell(row=row, column=2).style = "header_style"
        ws.merge_cells(f"B{row}:D{row}")
        row += 1

        cost_headers = ["Event Type", "Acute Cost", "Annual Follow-up"]
        for i, header in enumerate(cost_headers):
            ws.cell(row=row, column=i+2, value=header).style = "header_style"
        row += 1

        costs = self.extended_inputs.event_costs
        for event in EventType:
            acute, followup = costs.get_costs(event)
            ws.cell(row=row, column=2, value=event_labels[event])
            ws.cell(row=row, column=3, value=acute).number_format = f'"{currency}"#,##0'
            ws.cell(row=row, column=4, value=followup).number_format = f'"{currency}"#,##0'
            row += 1

        # Events avoided analysis (if results available)
        if self.extended_results and self.extended_results.event_results:
            row += 2
            ws.cell(row=row, column=2, value="EVENTS AVOIDED (5-YEAR)")
            ws.cell(row=row, column=2).style = "header_style"
            ws.merge_cells(f"B{row}:E{row}")
            row += 1

            ev_headers = ["Event Type", "Events Avoided", "Cost Avoided"]
            for i, header in enumerate(ev_headers):
                ws.cell(row=row, column=i+2, value=header).style = "header_style"
            row += 1

            event_results = self.extended_results.event_results
            for event in EventType:
                avoided = event_results.events_avoided.get(event, 0)
                cost_avoided = event_results.event_costs.get(event, 0)
                if avoided != 0 or cost_avoided != 0:
                    ws.cell(row=row, column=2, value=event_labels[event])
                    ws.cell(row=row, column=3, value=avoided).number_format = "#,##0"
                    ws.cell(row=row, column=4, value=cost_avoided).number_format = f'"{currency}"#,##0'
                    row += 1

            # Total
            ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
            ws.cell(row=row, column=3, value=sum(event_results.events_avoided.values())).style = "result_style"
            ws.cell(row=row, column=3).number_format = "#,##0"
            ws.cell(row=row, column=4, value=event_results.total_costs_avoided).style = "result_style"
            ws.cell(row=row, column=4).number_format = f'"{currency}"#,##0'

        # Set column widths
        ws.column_dimensions["B"].width = 30
        for col in range(3, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_psa_sheet(self):
        """Create PSA results sheet with distribution and histogram."""
        ws = self.wb.create_sheet("PSA Results")

        currency = self.extended_inputs.country.currency_symbol

        # Title
        ws.merge_cells("B2:H2")
        ws["B2"] = "PROBABILISTIC SENSITIVITY ANALYSIS RESULTS"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        if not self.extended_results or not self.extended_results.psa_results:
            ws["B4"] = "Run PSA using EnhancedBIMCalculator.run_probabilistic_sensitivity()"
            return

        psa = self.extended_results.psa_results

        # Summary statistics
        ws["B4"] = "SUMMARY STATISTICS"
        ws["B4"].style = "header_style"
        ws.merge_cells("B4:D4")

        stats = [
            ("Number of Iterations", psa.iterations, "#,##0"),
            ("Mean Budget Impact", psa.mean_impact, f'"{currency}"#,##0'),
            ("Median Budget Impact", psa.median_impact, f'"{currency}"#,##0'),
            ("Standard Deviation", psa.std_impact, f'"{currency}"#,##0'),
            (f"{psa.confidence_level:.0%} CI - Lower", psa.ci_lower, f'"{currency}"#,##0'),
            (f"{psa.confidence_level:.0%} CI - Upper", psa.ci_upper, f'"{currency}"#,##0'),
            ("Probability of Budget Increase", psa.prob_budget_increase, "0.0%"),
        ]

        for i, (label, value, fmt) in enumerate(stats):
            row = 5 + i
            ws.cell(row=row, column=2, value=label)
            cell = ws.cell(row=row, column=3, value=value)
            cell.number_format = fmt
            if "CI" in label or "Mean" in label:
                cell.style = "result_style"

        # Distribution data for histogram
        ws["B14"] = "DISTRIBUTION DATA (for histogram)"
        ws["B14"].style = "header_style"
        ws.merge_cells("B14:D14")

        # Create histogram bins
        import numpy as np
        if psa.impact_distribution:
            impacts = np.array(psa.impact_distribution)
            hist, bin_edges = np.histogram(impacts, bins=20)

            ws.cell(row=15, column=2, value="Bin Range").style = "header_style"
            ws.cell(row=15, column=3, value="Frequency").style = "header_style"

            for i, (count, edge) in enumerate(zip(hist, bin_edges[:-1])):
                row = 16 + i
                bin_label = f"{currency}{edge:,.0f} - {currency}{bin_edges[i+1]:,.0f}"
                ws.cell(row=row, column=2, value=bin_label)
                ws.cell(row=row, column=3, value=count)

            # Create bar chart as histogram
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Distribution of 5-Year Budget Impact"
            chart.y_axis.title = "Frequency"
            chart.x_axis.title = f"Budget Impact ({self.extended_inputs.costs.currency})"

            data = Reference(ws, min_col=3, min_row=15, max_col=3, max_row=15 + len(hist))
            cats = Reference(ws, min_col=2, min_row=16, max_col=2, max_row=15 + len(hist))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            chart.width = 15
            chart.height = 10
            ws.add_chart(chart, "E4")

        # Interpretation
        row = 40
        ws.cell(row=row, column=2, value="INTERPRETATION")
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)

        interpretations = [
            f"Based on {psa.iterations:,} Monte Carlo simulations:",
            f"- The mean 5-year budget impact is {currency}{psa.mean_impact:,.0f}",
            f"- There is a {psa.confidence_level:.0%} probability that the true impact falls between "
            f"{currency}{psa.ci_lower:,.0f} and {currency}{psa.ci_upper:,.0f}",
            f"- {psa.prob_budget_increase:.1%} of simulations resulted in a budget increase",
        ]

        for i, text in enumerate(interpretations):
            ws.cell(row=row + 1 + i, column=2, value=text)

        # Set column widths
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
